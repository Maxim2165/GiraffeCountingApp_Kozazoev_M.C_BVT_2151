from flask import Flask, request, jsonify, render_template, send_file
import cv2
import numpy as np
from ultralytics import YOLO
import sqlite3
from datetime import datetime
import json
import os
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font
import time

# Создаю экземпляр Flask для веб-сервера
app = Flask(__name__)

# Загружаю предобученную модель YOLOv8 для детекции объектов
model = YOLO('yolov8n.pt')

# Определяю путь к файлу базы данных SQLite
db_path = os.path.join(os.getcwd(), 'history.db')
# Подключаюсь к базе данных SQLite с поддержкой многопоточности
conn = sqlite3.connect(db_path, check_same_thread=False)
# Создаю курсор для выполнения SQL-запросов
cursor = conn.cursor()

# Создаю таблицу requests, если её нет, с нужными колонками
cursor.execute('''CREATE TABLE IF NOT EXISTS requests 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, result TEXT, processing_time REAL, filename TEXT)''')
# Проверяю структуру таблицы
cursor.execute("PRAGMA table_info(requests)")
# Получаю список колонок таблицы
columns = [col[1] for col in cursor.fetchall()]
# Добавляю колонку processing_time, если её нет
if 'processing_time' not in columns:
    cursor.execute("ALTER TABLE requests ADD COLUMN processing_time REAL")
# Добавляю колонку filename, если её нет
if 'filename' not in columns:
    cursor.execute("ALTER TABLE requests ADD COLUMN filename TEXT")
# Сохраняю изменения в базе
conn.commit()

# Проверяю наличие папки static и создаю её, если её нет
if not os.path.exists('static'):
    os.makedirs('static')

# Функция для генерации PDF-отчета с деталями
def generate_pdf_report(giraffe_count, timestamp, processing_time, filename, giraffe_details):
    # Определяю путь для сохранения PDF
    pdf_path = os.path.join('static', f'report_{timestamp}.pdf')
    # Создаю документ PDF с заданными размерами и отступами
    doc = SimpleDocTemplate(pdf_path, pagesize=letter, leftMargin=1.25*inch, rightMargin=1.25*inch, 
                           topMargin=1.25*inch, bottomMargin=1.25*inch)
    # Загружаю базовые стили
    styles = getSampleStyleSheet()
    # Определяю стиль для обычного текста
    custom_style = ParagraphStyle(name='CustomNormal', parent=styles['Normal'], fontSize=14, leading=21, alignment=0)
    # Определяю стиль для жирного текста
    custom_bold_style = ParagraphStyle(name='CustomBold', parent=styles['Normal'], fontSize=14, leading=21, fontName='Helvetica-Bold')
    # Определяю стиль для заголовка
    custom_heading = ParagraphStyle(name='CustomHeading', parent=styles['Heading1'], fontSize=20, alignment=1)
    # Указываю путь к обработанному изображению
    img_path = os.path.join('static', 'result.jpg')
    # Создаю список элементов для PDF
    elements = [Paragraph("Giraffe Detection Report", custom_heading),
                Paragraph(f"File: {filename}", custom_style),
                Paragraph(f"Date and Time: {timestamp}", custom_style),
                Paragraph(f"Processing Time: {processing_time:.2f} sec", custom_style)]
    # Если жирафы есть, добавляю их количество и координаты
    if giraffe_count > 0:
        elements.append(Paragraph(f"Number of Giraffes Detected: {giraffe_count}", custom_style))
        for i, (x, y, w, h) in enumerate(giraffe_details, 1):
            elements.append(Paragraph(f"Giraffe #{i}: Coordinates (x={x:.1f}, y={y:.1f}, width={w:.1f}, height={h:.1f})", custom_style))
    else:
        elements.append(Paragraph("No giraffes detected or invalid image format", custom_style))
    # Если изображение существует, добавляю его
    if os.path.exists(img_path):
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Processed Image", custom_bold_style))
        img = Image(img_path)
        img_width, img_height = img.drawWidth, img.drawHeight
        max_width, max_height = 400, 600
        if img_width > max_width or img_height > max_height:
            ratio = min(max_width / img_width, max_height / img_height)
            img.drawWidth = img_width * ratio
            img.drawHeight = img_height * ratio
        img.hAlign = 'CENTER'
        elements.append(img)
    # Собираю и сохраняю PDF
    doc.build(elements)
    return pdf_path

# Функция для генерации Excel-отчета с деталями
def generate_excel_report(giraffe_count, timestamp, processing_time, filename, giraffe_details):
    # Определяю путь для сохранения Excel
    excel_path = os.path.join('static', f'report_{timestamp}.xlsx')
    # Создаю новую книгу Excel
    wb = Workbook()
    # Активирую активный лист
    ws = wb.active
    # Назначаю название листа
    ws.title = "Giraffe Detection"
    # Объединяю ячейки для заголовка
    ws.merge_cells('A1:B1')
    # Заполняю заголовок
    ws['A1'] = "Giraffe Detection Report"
    # Делаю заголовок жирным
    ws['A1'].font = Font(bold=True)
    # Добавляю метку для имени файла
    ws['A2'] = "File"
    # Заполняю имя файла
    ws['B2'] = filename
    # Добавляю метку для даты
    ws['A3'] = "Date and Time"
    # Заполняю дату и время
    ws['B3'] = timestamp
    # Добавляю метку для времени обработки
    ws['A4'] = "Processing Time (sec)"
    # Заполняю время обработки
    ws['B4'] = round(processing_time, 2)
    # Если жирафы есть, добавляю их данные
    if giraffe_count > 0:
        ws['A5'] = "Number of Giraffes"
        ws['B5'] = giraffe_count
        for i, (x, y, w, h) in enumerate(giraffe_details, 1):
            ws[f"A{i+5}"] = f"Giraffe #{i}"
            ws[f"B{i+5}"] = f"x={x:.1f}, y={y:.1f}, width={w:.1f}, height={h:.1f}"
    else:
        ws['A5'] = "Status"
        ws['B5'] = "No giraffes detected or invalid image format"
    # Сохраняю файл Excel
    wb.save(excel_path)
    return excel_path

# Функция для получения истории обработок
def get_history():
    # Запрашиваю последние 5 записей из базы
    cursor.execute('SELECT timestamp, result, processing_time, filename FROM requests ORDER BY id DESC LIMIT 5')
    # Получаю все строки результата
    history = cursor.fetchall()
    # Создаю пустой список для обработанной истории
    parsed_history = []
    # Прохожу по каждой записи
    for item in history:
        timestamp, result_json, processing_time, filename = item
        result_data = json.loads(result_json)
        parsed_history.append((timestamp, result_data['count'], processing_time, filename))
    return parsed_history

# Определяю маршрут для главной страницы
@app.route('/')
def index():
    # Получаю историю обработок
    history = get_history()
    # Рендерю HTML-страницу с историей
    return render_template('index.html', history=history)

# Определяю маршрут для обработки изображения
@app.route('/process', methods=['POST'])
def process_image():
    # Проверяю наличие файла в запросе
    if 'image' not in request.files:
        return jsonify(error="No image uploaded"), 400
    # Получаю загруженный файл
    file = request.files['image']
    # Сохраняю имя файла
    filename = file.filename
    # Проверяю корректность расширения файла
    if not filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
        return jsonify(error="Invalid file format. Please upload an image (e.g., .png, .jpg, .jpeg)"), 400
    # Читаю данные файла
    img_data = file.read()
    # Проверяю, что файл не пустой
    if not img_data:
        return jsonify(error="Empty file uploaded"), 400
    # Декодирую изображение
    img = cv2.imdecode(np.frombuffer(img_data, np.uint8), cv2.IMREAD_COLOR)
    # Проверяю успешность декодирования
    if img is None:
        return jsonify(error="Failed to decode image. Please upload a valid image file"), 400
    # Записываю время начала обработки
    start_time = time.time()
    # Обрабатываю изображение моделью
    results = model(img)
    # Создаю визуализацию результатов
    output_img = results[0].plot()
    # Вычисляю время обработки
    processing_time = time.time() - start_time
    # Инициализирую счетчик жирафов
    giraffe_count = 0
    # Создаю список для деталей жирафов
    giraffe_details = []
    # Прохожу по всем обнаруженным объектам
    for box in results[0].boxes:
        class_id = int(box.cls[0])
        class_name = results[0].names[class_id]
        # Если объект - жираф, добавляю в счетчик и детали
        if class_name == "giraffe":
            giraffe_count += 1
            x, y, w, h = box.xywh[0].tolist()
            giraffe_details.append((x, y, w, h))
    # Сохраняю обработанное изображение
    cv2.imwrite(os.path.join('static', 'result.jpg'), output_img, [int(cv2.IMWRITE_JPEG_QUALITY), 90])
    # Форматирую текущую дату и время
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    # Подготавливаю данные для записи в базу
    result_data = {'count': giraffe_count, 'classes': results[0].names}
    # Добавляю запись в базу данных
    cursor.execute('INSERT INTO requests (timestamp, result, processing_time, filename) VALUES (?, ?, ?, ?)', 
                   (timestamp, json.dumps(result_data), processing_time, filename))
    # Сохраняю изменения в базе
    conn.commit()
    # Генерирую PDF-отчет
    pdf_path = generate_pdf_report(giraffe_count, timestamp, processing_time, filename, giraffe_details)
    # Генерирую Excel-отчет
    excel_path = generate_excel_report(giraffe_count, timestamp, processing_time, filename, giraffe_details)
    # Формирую URL для обработанного изображения с временной меткой для предотвращения кеширования
    image_url = f"/static/result.jpg?{timestamp}"
    # Возвращаю результат в JSON с URL изображения
    return jsonify({'count': giraffe_count, 'timestamp': timestamp, 'image_url': image_url})

# Определяю маршрут для получения истории
@app.route('/get_history')
def get_history_route():
    # Получаю историю обработок
    history = get_history()
    # Возвращаю историю в формате JSON
    return jsonify(history)

# Определяю маршрут для очистки истории
@app.route('/clear_history', methods=['POST'])
def clear_history():
    # Удаляю все записи из таблицы
    cursor.execute('DELETE FROM requests')
    # Сохраняю изменения
    conn.commit()
    # Возвращаю успешный статус
    return jsonify({'status': 'success'})

# Определяю маршрут для скачивания отчета
@app.route('/download_report/<report_type>')
def download_report(report_type):
    # Получаю timestamp из запроса
    timestamp = request.args.get('timestamp', '')
    # Если timestamp не указан, беру последнюю запись
    if not timestamp:
        cursor.execute('SELECT timestamp, filename FROM requests ORDER BY id DESC LIMIT 1')
    # Иначе ищу запись по timestamp
    else:
        cursor.execute('SELECT timestamp, filename FROM requests WHERE timestamp = ?', (timestamp,))
    # Получаю результат
    latest = cursor.fetchone()
    # Проверяю наличие записей
    if not latest:
        return "No reports available", 404
    # Распаковываю данные
    timestamp, filename = latest
    # Если запрашивают PDF
    if report_type == 'pdf':
        file_path = os.path.join('static', f'report_{timestamp}.pdf')
        # Проверяю существование файла и отправляю его
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=f'report_{filename}.pdf')
        return "PDF not found", 404
    # Если запрашивают Excel
    elif report_type == 'excel':
        file_path = os.path.join('static', f'report_{timestamp}.xlsx')
        # Проверяю существование файла и отправляю его
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=f'report_{filename}.xlsx')
        return "Excel not found", 404
    # Ошибка при неверном типе отчета
    return "Invalid report type", 400

if __name__ == '__main__':
    # Запускаю сервер на всех интерфейсах с отладкой
    app.run(host='0.0.0.0', port=5000, debug=True)